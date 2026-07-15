import os
import re
import json
import requests
import psycopg2
import urllib.parse
import socket
from datetime import datetime, timedelta, date, timezone
from http.server import HTTPServer, BaseHTTPRequestHandler
from io import BytesIO

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")
WEBHOOK_URL = os.getenv("WEBHOOK_URL")
PORT = int(os.getenv("PORT", 8080))
OWNER_ID = 700080578  # Only this user can use owner-only commands

SITES = {
    "assamtenders": {
        "display": "Assamtenders",
        "url": "https://assamtenders.gov.in/nicgep/app"
    },
    "etenders": {
        "display": "Etenders",
        "url": "https://etenders.gov.in/eprocure/app"
    },
    "pmgsy": {
        "display": "PMGSY Assam",
        "url": "https://pmgsytendersasm.gov.in/nicgep/app?page=Home&service=page"
    },
    "ongc": {
        "display": "ONGC",
        "url": "https://tenders.ongc.co.in/web/tendersweb"
    },
}

WATCHLIST = [
    "Jorhat", "Sivasagar", "Sibsagar", "Charaideo", "Nazira",
    "Sonari", "Amguri", "Demow", "Lakwa",
    "Simaluguri", "Moran", "Duliajan", "Tinsukia",
    "Dibrugarh", "Golaghat", "Mariani", "Bhojo",
    "Assam"
]

# In-memory cache for pending export requests (per chat_id)
# Stores parsed source/date range while waiting for format button tap
PENDING_EXPORTS = {}


# ── Database ──────────────────────────────────────────────────────────────────

def get_conn():
    url = DATABASE_URL.strip()
    parsed = urllib.parse.urlparse(url)
    ipv4 = socket.getaddrinfo(parsed.hostname, parsed.port or 5432, socket.AF_INET)[0][4][0]
    return psycopg2.connect(
        host=ipv4,
        port=parsed.port or 5432,
        dbname=parsed.path.lstrip("/"),
        user=parsed.username,
        password=parsed.password,
        sslmode="require",
        connect_timeout=10
    )


def query_db(sql, params=()):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return rows
    except Exception as e:
        print(f"DB query error: {type(e).__name__}: {e}")
        return []


def execute_db(sql, params=()):
    """For INSERT/DELETE statements. Returns True on success."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(sql, params)
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f"DB execute error: {type(e).__name__}: {e}")
        return False


# ── Telegram ──────────────────────────────────────────────────────────────────

def send(chat_id, text, reply_markup=None):
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "HTML"
    }
    if reply_markup:
        payload["reply_markup"] = json.dumps(reply_markup)
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json=payload,
            timeout=30
        )
    except Exception as e:
        print(f"Send error: {e}")


def send_document(chat_id, filename, file_bytes, caption=""):
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
            data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
            files={"document": (filename, file_bytes)},
            timeout=60
        )
    except Exception as e:
        print(f"Send document error: {e}")


def answer_callback(callback_query_id, text=""):
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/answerCallbackQuery",
            json={"callback_query_id": callback_query_id, "text": text},
            timeout=15
        )
    except Exception as e:
        print(f"Answer callback error: {e}")


def edit_message(chat_id, message_id, text, reply_markup=None):
    payload = {
        "chat_id": chat_id,
        "message_id": message_id,
        "text": text,
        "parse_mode": "HTML"
    }
    if reply_markup:
        payload["reply_markup"] = json.dumps(reply_markup)
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/editMessageText",
            json=payload,
            timeout=30
        )
    except Exception as e:
        print(f"Edit message error: {e}")


def set_webhook():
    if not WEBHOOK_URL:
        print("WEBHOOK_URL not set — skipping")
        return
    resp = requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/setWebhook",
        json={"url": f"{WEBHOOK_URL}/webhook"}
    )
    print(f"Webhook set: {resp.json()}")


# ── Formatting ────────────────────────────────────────────────────────────────

def truncate(text, length=75):
    return text if len(text) <= length else text[:length].rstrip() + "..."


def get_matched_location(title):
    for place in WATCHLIST:
        if place.lower() in title.lower():
            return place
    return ""


def format_row(row, idx=None):
    _, title, ref, closing, opening, date_found, source = row
    site = SITES.get(source, {})
    display = site.get("display", source)
    prefix = f"{idx}. " if idx else "🔹 "
    closing_date = closing.split()[0] if closing else "N/A"
    location = get_matched_location(title)
    result = f"{prefix}<b>{truncate(title)}</b>\n"
    if location:
        result += f"📍 {location}\n"
    result += f"📎 {ref}\n"
    result += f"📅 {closing_date}\n"
    result += f"🏢 {display}\n"
    return result


def send_results(chat_id, rows, header, reply_markup=None):
    if not rows:
        send(chat_id, f"No results found for: {header}")
        return
    reply = f"<b>{header}</b>\n({len(rows)} found)\n\n"
    for idx, row in enumerate(rows, 1):
        reply += format_row(row, idx) + "\n"

    if len(reply) <= 4096:
        send(chat_id, reply, reply_markup)
    else:
        chunks = [reply[i:i+4096] for i in range(0, len(reply), 4096)]
        for i, chunk in enumerate(chunks):
            # Only attach buttons to the last chunk
            send(chat_id, chunk, reply_markup if i == len(chunks) - 1 else None)


# ── Existing Commands ─────────────────────────────────────────────────────────

def cmd_today(chat_id):
    today = date.today()
    rows = query_db(
        "SELECT type, title, ref, closing, opening, date_found, source FROM alerts WHERE date_found = %s ORDER BY id DESC",
        (today,)
    )
    send_results(chat_id, rows, f"📋 TENDERS LISTED TODAY ({today})")


def cmd_week(chat_id):
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    rows = query_db(
        "SELECT type, title, ref, closing, opening, date_found, source FROM alerts WHERE date_found >= %s ORDER BY date_found DESC, id DESC",
        (week_start,)
    )
    send_results(chat_id, rows, f"📋 TENDERS LISTED THIS WEEK ({week_start} → {today})")


def cmd_closing_today(chat_id):
    today_str = datetime.now().strftime("%d-%b-%Y").upper()
    rows = query_db(
        "SELECT type, title, ref, closing, opening, date_found, source FROM alerts WHERE UPPER(closing) LIKE %s ORDER BY id DESC",
        (f"%{today_str}%",)
    )
    send_results(chat_id, rows, f"⏰ TENDERS CLOSING TODAY ({today_str})")


def cmd_closing_week(chat_id):
    results = []
    rows = query_db(
        "SELECT type, title, ref, closing, opening, date_found, source FROM alerts ORDER BY id DESC"
    )
    now = datetime.now()
    for row in rows:
        closing_str = row[3].split()[0] if row[3] else ""
        try:
            closing_date = datetime.strptime(closing_str, "%d-%b-%Y")
            if now <= closing_date <= now + timedelta(days=7):
                results.append(row)
        except:
            pass
    send_results(chat_id, results, "⏰ TENDERS CLOSING THIS WEEK")


def cmd_latest(chat_id):
    rows = query_db(
        "SELECT type, title, ref, closing, opening, date_found, source FROM alerts ORDER BY id DESC LIMIT 5"
    )
    send_results(chat_id, rows, "🔴 LATEST 5 TENDERS")


def cmd_search(chat_id, keyword):
    if not keyword:
        send(chat_id, "Please provide a keyword.\nExample: /search roads")
        return
    rows = query_db(
        "SELECT type, title, ref, closing, opening, date_found, source FROM alerts WHERE LOWER(title) LIKE %s ORDER BY id DESC",
        (f"%{keyword.lower()}%",)
    )
    send_results(chat_id, rows, f"🔍 SEARCH: '{keyword}'")


def cmd_source(chat_id, source):
    if source not in SITES:
        send(chat_id, "Unknown source. Use:\n/source assamtenders\n/source etenders\n/source pmgsy\n/source ongc")
        return
    rows = query_db(
        "SELECT type, title, ref, closing, opening, date_found, source FROM alerts WHERE source = %s ORDER BY id DESC LIMIT 10",
        (source,)
    )
    display = SITES[source]["display"]
    send_results(chat_id, rows, f"📍 LATEST FROM {display}")


def cmd_stats(chat_id):
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    total_row = query_db("SELECT COUNT(*) FROM alerts")
    total = total_row[0][0] if total_row else 0
    today_count = query_db("SELECT COUNT(*) FROM alerts WHERE date_found = %s", (today,))[0][0]
    week_count = query_db("SELECT COUNT(*) FROM alerts WHERE date_found >= %s", (week_start,))[0][0]
    month_count = query_db("SELECT COUNT(*) FROM alerts WHERE date_found >= %s", (month_start,))[0][0]

    assam_count = query_db("SELECT COUNT(*) FROM alerts WHERE source = 'assamtenders'")[0][0]
    etenders_count = query_db("SELECT COUNT(*) FROM alerts WHERE source = 'etenders'")[0][0]
    pmgsy_count = query_db("SELECT COUNT(*) FROM alerts WHERE source = 'pmgsy'")[0][0]
    ongc_count = query_db("SELECT COUNT(*) FROM alerts WHERE source = 'ongc'")[0][0]

    latest = query_db("SELECT title, date_found FROM alerts ORDER BY id DESC LIMIT 1")

    msg = (
        "📊 <b>BOT STATISTICS</b>\n"
        "━━━━━━━━━━━━━━━━━━\n\n"
        f"📌 <b>Total tenders tracked:</b> {total}\n"
        f"📅 <b>Today:</b> {today_count}\n"
        f"📆 <b>This week:</b> {week_count}\n"
        f"🗓 <b>This month:</b> {month_count}\n\n"
        f"📍 <b>By source:</b>\n"
        f"  • Assamtenders: {assam_count}\n"
        f"  • Etenders: {etenders_count}\n"
        f"  • PMGSY Assam: {pmgsy_count}\n"
        f"  • ONGC: {ongc_count}\n"
    )

    if latest:
        msg += f"\n🕐 <b>Last found:</b>\n{latest[0][0][:50]}...\n({latest[0][1]})"

    send(chat_id, msg)


def cmd_ping(chat_id):
    IST = timezone(timedelta(hours=5, minutes=30))
    now = datetime.now(IST).strftime("%d %b %Y %I:%M %p IST")

    last_run_row = query_db("SELECT value FROM bot_meta WHERE key = 'last_run'")
    last_run = last_run_row[0][0] if last_run_row else "Unknown"

    today = date.today()
    today_count_row = query_db("SELECT COUNT(*) FROM alerts WHERE date_found = %s", (today,))
    today_count = today_count_row[0][0] if today_count_row else 0
    total_row = query_db("SELECT COUNT(*) FROM alerts")
    total = total_row[0][0] if total_row else 0

    send(chat_id,
        f"🟢 <b>BOT STATUS</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🕐 <b>Current time:</b> {now}\n"
        f"📡 <b>Monitor last run:</b> {last_run}\n"
        f"🗄 <b>New today:</b> {today_count}\n"
        f"📊 <b>Total tracked:</b> {total}\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"✅ Everything running fine"
    )


def cmd_status(chat_id):
    today = date.today()

    sources = ["assamtenders", "etenders", "pmgsy", "ongc"]
    status_lines = ""

    for source in sources:
        rows = query_db(
            "SELECT date_found FROM alerts WHERE source = %s ORDER BY id DESC LIMIT 1",
            (source,)
        )
        if rows:
            last_seen = rows[0][0]
            days_ago = (today - last_seen).days
            if days_ago == 0:
                label = "today ✅"
            elif days_ago == 1:
                label = "yesterday ⚠️"
            else:
                label = f"{days_ago} days ago ❌"
            status_lines += f"  • {source}: last entry {label}\n"
        else:
            status_lines += f"  • {source}: no data yet\n"

    total = query_db("SELECT COUNT(*) FROM alerts")[0][0]
    week_start = today - timedelta(days=today.weekday())
    this_week = query_db(
        "SELECT COUNT(*) FROM alerts WHERE date_found >= %s", (week_start,)
    )[0][0]

    send(chat_id,
        f"📡 <b>BOT STATUS</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n\n"
        f"🗄 <b>Database:</b>\n"
        f"  • Total tenders: {total}\n"
        f"  • This week: {this_week}\n\n"
        f"🌐 <b>Sources:</b>\n"
        f"{status_lines}\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"✅ Bot is running normally"
    )


# ── Save / Saved Commands ─────────────────────────────────────────────────────

def cmd_save(chat_id, user_id, ref):
    if not ref:
        send(chat_id, "Please provide a tender reference to save.\nExample: /save AS21689")
        return

    rows = query_db(
        "SELECT title, ref, closing, source FROM alerts WHERE ref = %s ORDER BY id DESC LIMIT 1",
        (ref,)
    )
    if not rows:
        send(chat_id, f"No tender found with reference '{ref}'.\nCheck the 📎 Ref or 🛣 Road Code in the alert message.")
        return

    title, found_ref, closing, source = rows[0]
    success = execute_db(
        """INSERT INTO saved_tenders (user_id, tender_ref, title, source, closing)
           VALUES (%s, %s, %s, %s, %s)
           ON CONFLICT (user_id, tender_ref) DO NOTHING""",
        (user_id, found_ref, title, source, closing)
    )
    if success:
        send(chat_id, f"⭐ <b>Saved!</b>\n\n{truncate(title)}\n📎 {found_ref}")
    else:
        send(chat_id, "Failed to save tender. Try again later.")


def cmd_saved(chat_id, user_id):
    rows = query_db(
        "SELECT title, tender_ref, source, closing FROM saved_tenders WHERE user_id = %s ORDER BY saved_at DESC",
        (user_id,)
    )
    if not rows:
        send(chat_id, "⭐ You have no saved tenders yet.\nUse /save [ref] to bookmark one.")
        return

    reply = f"⭐ <b>SAVED TENDERS</b> ({len(rows)})\n\n"
    for idx, (title, ref, source, closing) in enumerate(rows, 1):
        display = SITES.get(source, {}).get("display", source)
        closing_date = closing.split()[0] if closing else "N/A"
        reply += f"{idx}. <b>{truncate(title)}</b>\n"
        reply += f"📎 {ref}\n"
        reply += f"📅 {closing_date}\n"
        reply += f"🏢 {display}\n\n"

    reply_markup = {
        "inline_keyboard": [[
            {"text": "📤 Export Saved", "callback_data": "export_saved"}
        ]]
    }

    if len(reply) <= 4096:
        send(chat_id, reply, reply_markup)
    else:
        chunks = [reply[i:i+4096] for i in range(0, len(reply), 4096)]
        for i, chunk in enumerate(chunks):
            send(chat_id, chunk, reply_markup if i == len(chunks) - 1 else None)


def cmd_unsave(chat_id, user_id, ref):
    if not ref:
        send(chat_id, "Please provide a tender reference to remove.\nExample: /unsave AS21689")
        return
    execute_db(
        "DELETE FROM saved_tenders WHERE user_id = %s AND tender_ref = %s",
        (user_id, ref)
    )
    send(chat_id, f"🗑 Removed from saved tenders: {ref}")


# ── Export ────────────────────────────────────────────────────────────────────

MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}


def parse_export_args(text):
    """
    Parse '/export [source] [date1] [date2]' flexibly.
    Returns (source_or_None, start_date, end_date, used_default_range)
    """
    parts = text.replace("/export", "").strip().split()
    source = None
    dates = []

    for part in parts:
        if part.lower() in SITES:
            source = part.lower()
        else:
            # Try to parse as date in DD/MM/YYYY or DD-MM-YYYY
            for sep in ["/", "-"]:
                if sep in part:
                    try:
                        d, m, y = part.split(sep)
                        dates.append(date(int(y), int(m), int(d)))
                        break
                    except (ValueError, IndexError):
                        pass

    used_default = False
    if len(dates) >= 2:
        start_date, end_date = sorted(dates[:2])
    elif len(dates) == 1:
        start_date = dates[0]
        end_date = date.today()
    else:
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
        used_default = True

    return source, start_date, end_date, used_default


def cmd_export(chat_id, text):
    source, start_date, end_date, used_default = parse_export_args(text)

    source_display = SITES.get(source, {}).get("display", "All Sources") if source else "All Sources"
    range_label = "Last 30 days (default)" if used_default else f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

    PENDING_EXPORTS[chat_id] = {
        "source": source,
        "start_date": start_date,
        "end_date": end_date,
        "used_default": used_default,
        "scope": "main"
    }

    reply_markup = {
        "inline_keyboard": [[
            {"text": "📊 Excel", "callback_data": "fmt_excel"},
            {"text": "📄 PDF", "callback_data": "fmt_pdf"},
            {"text": "📋 CSV", "callback_data": "fmt_csv"}
        ]]
    }

    send(chat_id,
        f"📤 <b>Export Tenders</b>\n"
        f"Source: {source_display}\n"
        f"Range: {range_label}\n\n"
        f"Choose a format:",
        reply_markup
    )


def get_export_rows(pending):
    if pending["scope"] == "saved":
        rows = query_db(
            "SELECT title, tender_ref, source, closing, saved_at FROM saved_tenders WHERE user_id = %s ORDER BY saved_at DESC",
            (pending["user_id"],)
        )
        # saved_tenders table has no location column — fall back to title guess
        return [
            (r[0], r[1], r[2], r[3] or "", get_matched_location(r[0]), str(r[4])[:10])
            for r in rows
        ]
    else:
        sql = "SELECT title, ref, source, closing, date_found, location FROM alerts WHERE date_found BETWEEN %s AND %s"
        params = [pending["start_date"], pending["end_date"]]
        if pending["source"]:
            sql += " AND source = %s"
            params.append(pending["source"])
        sql += " ORDER BY date_found DESC, id DESC"
        rows = query_db(sql, tuple(params))
        return [
            (r[0], r[1], r[2], r[3] or "", r[5] or get_matched_location(r[0]), str(r[4]))
            for r in rows
        ]


def generate_csv(rows, meta_note):
    import csv
    output = BytesIO()
    text_buffer = []
    text_buffer.append(["Title", "Reference", "Source", "Closing", "Location", "Date Found"])
    for title, ref, source, closing, location, date_found in rows:
        display = SITES.get(source, {}).get("display", source)
        text_buffer.append([title, ref, display, closing, location, date_found])

    import io
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow([meta_note])
    writer.writerow([])
    for row in text_buffer:
        writer.writerow(row)

    output.write(sio.getvalue().encode("utf-8"))
    output.seek(0)
    return output


def generate_excel(rows, meta_note):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenders"

    ws.merge_cells("A1:F1")
    ws["A1"] = meta_note
    ws["A1"].font = Font(italic=True, size=10, color="666666")

    headers = ["Title", "Reference", "Source", "Closing", "Location", "Date Found"]
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for i, (title, ref, source, closing, location, date_found) in enumerate(rows, header_row + 1):
        display = SITES.get(source, {}).get("display", source)
        ws.cell(row=i, column=1, value=title)
        ws.cell(row=i, column=2, value=ref)
        ws.cell(row=i, column=3, value=display)
        ws.cell(row=i, column=4, value=closing)
        ws.cell(row=i, column=5, value=location)
        ws.cell(row=i, column=6, value=date_found)

    widths = [50, 25, 18, 20, 12, 14]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf(rows, meta_note):
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", size=8)

    pdf.set_font("Helvetica", "I", 9)
    pdf.cell(0, 6, meta_note, ln=True)
    pdf.ln(2)

    headers = ["Title", "Reference", "Source", "Closing", "Location", "Date Found"]
    widths = [95, 40, 28, 32, 40, 25]
    row_height = 7

    def fit_text(pdf_obj, text, width, pad=2):
        """Truncate text with '...' so it fits within the given column width."""
        text = str(text)
        max_width = width - pad
        if pdf_obj.get_string_width(text) <= max_width:
            return text
        while text and pdf_obj.get_string_width(text + "...") > max_width:
            text = text[:-1]
        return text + "..." if text else ""

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(68, 114, 196)
    pdf.set_text_color(255, 255, 255)
    for h, w in zip(headers, widths):
        pdf.cell(w, row_height, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", size=7)
    pdf.set_text_color(0, 0, 0)
    fill = False
    for title, ref, source, closing, location, date_found in rows:
        display = SITES.get(source, {}).get("display", source)
        pdf.set_fill_color(240, 240, 240)

        values = [
            str(title).encode('latin-1', 'replace').decode('latin-1'),
            str(ref),
            str(display),
            str(closing),
            str(location).encode('latin-1', 'replace').decode('latin-1'),
            str(date_found),
        ]

        for val, w in zip(values, widths):
            safe_val = fit_text(pdf, val, w)
            pdf.cell(w, row_height, safe_val, border=1, fill=fill, align="L")
        pdf.ln()
        fill = not fill

    output = BytesIO()
    pdf_bytes = pdf.output(dest="S")
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode("latin-1")
    output.write(pdf_bytes)
    output.seek(0)
    return output


def handle_export_format(chat_id, fmt, pending):
    rows = get_export_rows(pending)

    if not rows:
        send(chat_id, "No tenders found for this export.")
        return

    if pending["scope"] == "saved":
        meta_note = f"Exported: {datetime.now().strftime('%d %b %Y %H:%M')} | Saved Tenders | Total: {len(rows)}"
        filename_base = "saved_tenders"
    else:
        range_label = "Last 30 days (auto)" if pending.get("used_default") else f"{pending['start_date'].strftime('%d-%m-%Y')} to {pending['end_date'].strftime('%d-%m-%Y')}"
        source_label = SITES.get(pending["source"], {}).get("display", "All Sources") if pending.get("source") else "All Sources"
        meta_note = f"Exported: {datetime.now().strftime('%d %b %Y %H:%M')} | Range: {range_label} | Source: {source_label} | Total: {len(rows)}"
        filename_base = f"tenders_{pending['start_date'].strftime('%d-%m-%Y')}_to_{pending['end_date'].strftime('%d-%m-%Y')}"

    try:
        if fmt == "excel":
            file_bytes = generate_excel(rows, meta_note)
            send_document(chat_id, f"{filename_base}.xlsx", file_bytes, f"📊 Excel export ready — {len(rows)} tenders")
        elif fmt == "pdf":
            file_bytes = generate_pdf(rows, meta_note)
            send_document(chat_id, f"{filename_base}.pdf", file_bytes, f"📄 PDF export ready — {len(rows)} tenders")
        elif fmt == "csv":
            file_bytes = generate_csv(rows, meta_note)
            send_document(chat_id, f"{filename_base}.csv", file_bytes, f"📋 CSV export ready — {len(rows)} tenders")
    except Exception as e:
        print(f"Export generation error: {e}")
        send(chat_id, f"⚠️ Export failed: {e}")


def cmd_help(chat_id):
    reply = (
        "🤖 <b>TENDER MONITOR BOT</b>\n\n"
        "📋 <b>LISTING COMMANDS:</b>\n"
        "/today — Tenders listed today\n"
        "/week — Tenders listed this week\n"
        "/latest — Last 5 tenders\n\n"
        "⏰ <b>CLOSING COMMANDS:</b>\n"
        "/closing_today — Closing today\n"
        "/closing_week — Closing this week\n\n"
        "🔍 <b>FILTER COMMANDS:</b>\n"
        "/search [keyword] — Search by keyword\n"
        "/source [name] — Filter by site\n"
        "   assamtenders, etenders, pmgsy, ongc\n\n"
        "⭐ <b>SAVE COMMANDS:</b>\n"
        "/save [ref] — Bookmark a tender\n"
        "/saved — View saved tenders\n"
        "/unsave [ref] — Remove a bookmark\n\n"
        "📤 <b>EXPORT:</b>\n"
        "/export — Last 30 days, all sources\n"
        "/export 10/06/2026 15/06/2026\n"
        "/export pmgsy — Specific source\n"
        "/export ongc 01/06/2026 30/06/2026\n\n"
        "📊 <b>OTHER:</b>\n"
        "/stats — Bot statistics\n"
        "/help — Show this message"
    )
    send(chat_id, reply)


# ── Update Handler ────────────────────────────────────────────────────────────

def handle_message(msg):
    chat_id = msg.get("chat", {}).get("id")
    user_id = msg.get("from", {}).get("id")
    text = msg.get("text", "").strip()

    if not chat_id or not text:
        return

    print(f"Command: {text} from user_id: {user_id}")

    if text.startswith("/ping") or text.startswith("/status"):
        if user_id == OWNER_ID:
            if text.startswith("/ping"):
                cmd_ping(chat_id)
            else:
                cmd_status(chat_id)
        return

    if text.startswith("/today"):
        cmd_today(chat_id)
    elif text.startswith("/week"):
        cmd_week(chat_id)
    elif text.startswith("/closing_today"):
        cmd_closing_today(chat_id)
    elif text.startswith("/closing_week"):
        cmd_closing_week(chat_id)
    elif text.startswith("/latest"):
        cmd_latest(chat_id)
    elif text.startswith("/unsave"):
        ref = text.replace("/unsave", "").strip()
        cmd_unsave(chat_id, user_id, ref)
    elif text.startswith("/saved"):
        cmd_saved(chat_id, user_id)
    elif text.startswith("/save"):
        ref = text.replace("/save", "").strip()
        cmd_save(chat_id, user_id, ref)
    elif text.startswith("/export"):
        cmd_export(chat_id, text)
    elif text.startswith("/search"):
        keyword = text.replace("/search", "").strip()
        cmd_search(chat_id, keyword)
    elif text.startswith("/source"):
        source = text.replace("/source", "").strip()
        cmd_source(chat_id, source)
    elif text.startswith("/stats"):
        cmd_stats(chat_id)
    elif text.startswith("/help") or text.startswith("/start"):
        cmd_help(chat_id)
    else:
        send(chat_id, "Unknown command. Type /help to see available commands.")


def handle_callback(callback_query):
    callback_id = callback_query.get("id")
    data = callback_query.get("data", "")
    chat_id = callback_query.get("message", {}).get("chat", {}).get("id")
    message_id = callback_query.get("message", {}).get("message_id")
    user_id = callback_query.get("from", {}).get("id")

    if not chat_id:
        return

    print(f"Callback: {data} from user_id: {user_id}")

    if data == "export_saved":
        PENDING_EXPORTS[chat_id] = {"scope": "saved", "user_id": user_id}
        reply_markup = {
            "inline_keyboard": [[
                {"text": "📊 Excel", "callback_data": "fmt_excel"},
                {"text": "📄 PDF", "callback_data": "fmt_pdf"},
                {"text": "📋 CSV", "callback_data": "fmt_csv"}
            ]]
        }
        answer_callback(callback_id)
        send(chat_id, "📤 <b>Export Saved Tenders</b>\nChoose a format:", reply_markup)
        return

    if data.startswith("fmt_"):
        fmt = data.replace("fmt_", "")
        pending = PENDING_EXPORTS.get(chat_id)
        answer_callback(callback_id, "Generating export...")
        if not pending:
            send(chat_id, "Export request expired. Please run /export again.")
            return
        if pending.get("scope") == "saved" and "user_id" not in pending:
            pending["user_id"] = user_id
        handle_export_format(chat_id, fmt, pending)
        return

    answer_callback(callback_id)


def handle(update):
    if "message" in update:
        handle_message(update["message"])
    elif "callback_query" in update:
        handle_callback(update["callback_query"])


# ── Webhook Server ────────────────────────────────────────────────────────────

class WebhookHandler(BaseHTTPRequestHandler):

    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Tender Bot is running!")

    def do_POST(self):
        if self.path == "/webhook":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            try:
                update = json.loads(body)
                handle(update)
            except Exception as e:
                print(f"Webhook error: {e}")
            self.send_response(200)
            self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()

    def log_message(self, format, *args):
        pass


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Testing database connection...")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM alerts")
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        print(f"Database connected! {count} rows in alerts table.")
    except Exception as e:
        print(f"DATABASE CONNECTION FAILED: {type(e).__name__}: {e}")

    set_webhook()
    print(f"Starting webhook server on port {PORT}...")
    server = HTTPServer(("0.0.0.0", PORT), WebhookHandler)
    server.serve_forever()


if __name__ == "__main__":
    main()
